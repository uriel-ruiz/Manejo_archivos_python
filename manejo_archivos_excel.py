from openpyxl import Workbook, load_workbook

# Crear un archivo Excel y escribir datos
wb = Workbook()
sheet = wb.active
sheet['A1'] = 'Nombre'
sheet['B1'] = 'Edad'
sheet.append(['Ana', 25])
sheet.append(['Luis', 30])
wb.save("archivo.xlsx")

# Leer el archivo Excel
wb = load_workbook("archivo.xlsx")
sheet = wb.active
print("Contenido del archivo Excel:")
for row in sheet.iter_rows(values_only=True):
    print(row)
